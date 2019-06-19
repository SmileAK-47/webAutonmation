# encoding = utf -8
from openpyxl import load_workbook


class ParseExcel(object):

    def __init__(self, excelPath, sheetName):
        # 将读取到的Excel加载到内存
        self.wb = load_workbook(excelPath)
        # 通过工作表名称获取一个工作表对象
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        # 获取工作表中醉倒的行号
        self.maxRowNum = self.sheet.max_row

    def getDateaFromSheet(self):
        # 用于存储从工作表中读取出来的数据
        dataList = []
        hah = self.sheet.rows
        # 因为工作表中第一行是标题，所以需要去掉
        for line in list(hah)[1:]:
            # 遍历工作表中数据区域的每一行
            # 并将每一行中各个单元格的数据存取与列表tmpList中
            # 然后再将存放一行的数据的列表加到最终数据列表dataList中
            tmpList = []
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            dataList.append(tmpList)

        # 将获取工作表中的所有数据的迭代对象返回
        return list(dataList)


if __name__ == '__main__':
    excelPath = r'G:\KeyWordFromeWork\testData\NBA.xlsx'
    sheetNmae = "搜索数据"
    pe = ParseExcel(excelPath, sheetNmae)
    for i in pe.getDateaFromSheet():
        print(i[0], i[1])
